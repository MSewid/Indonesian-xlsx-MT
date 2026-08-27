# =========================
# INSTALL DEPENDENCIES
# =========================

import re
from pathlib import Path
from html import unescape

import openpyxl
import argostranslate.package
import argostranslate.translate
from tqdm import tqdm

# =========================
# SETTINGS
# =========================

from pathlib import Path

BASE_DIR = Path(__file__).parent

INPUT_FILE = BASE_DIR / "TPAY Shahid SMS Template.xlsx"
OUTPUT_FILE = BASE_DIR / "TPAY Shahid SMS Template_ID.xlsx"

FROM_LANG = "en"
TO_LANG = "id"

# Skip the first worksheet only
SKIP_FIRST_SHEET = False

# Headers are expected in row 1
HEADER_ROW = 1

# If True, overwrite existing translations in AI Sub / AI Subs columns.
# If False, only fill empty AI Sub / AI Subs cells.
OVERWRITE_EXISTING_AI_SUB = True

SKIP_FORMULAS = True

# Source columns to translate, matched by first cell/header in row 1
SOURCE_HEADERS_TO_TRANSLATE = {
    "displaytext",
    "keywords",
    "metatitle",
    "screenname",
    "title",
    "metadesciption",      # typo as provided
    "metadescription",     # correct spelling also supported
    "iconname",
    "name enlgish",        # typo as provided
    "name english",        # correct spelling also supported
    "name (enlgish)",      # typo as provided
    "name (english)",      # correct spelling also supported
    "English",
    "English (EN)",
    "EN",
    "msg_en",
    "en",
    "English Copy",
    "Long Synopsis EN",
    "Short Synopsis EN",
    "Supplier Synopsis EN",
    "Long Synopsis VOD Override EN",
    "Short Synopsis VOD Override EN",
    "Supplier Synopsis VOD Override EN",
    "English Body",
    "EN Template SMS",
    "Welcome SMS",
}

# Target columns to write into, matched by first cell/header in row 1
TARGET_AI_HEADERS = {
    "ai sub",
    "ai subs",
    "ai_sub",
    "ai_subs",
    "ai",
    "IN",
    "AI Sub",
    "AI (Indonesian)",
}

# =========================
# SOURCE TYPO FIXES
# =========================

SOURCE_FIXES = {
    "Janurary": "January",
    "Subscrption": "Subscription",
    "Subscraption": "Subscription",
    "Mangement": "Management",
    "Budle": "Bundle",
    "Origials": "Originals",
    "excluisve": "exclusive",
    "eries": "series",
    "untill": "until",
    "Do yu neeed": "Do you need",
    "This filed is required": "This field is required",
    "langaue": "language",
    "pleasec": "please",
    "aspecial": "special",
    "Kepp": "Keep",
    "Fantaasy": "Fantasy",
    "Tittles": "Titles",
    "ldoesnt": "does not",
}

# =========================
# MANUAL EXACT TRANSLATIONS
# Add frequent short UI strings here for better quality.
# =========================

MANUAL_TRANSLATIONS = {
    "Continue": "Lanjutkan",
    "Cancel": "Batal",
    "OK": "OKE",
    "Ok": "Oke",
    "View packs": "Lihat paket",
    "Verify": "Verifikasi",
    "Resend code": "Kirim ulang kode",
    "Create account": "Buat akun",
    "Agree": "Setuju",
    "Disagree": "Tidak setuju",
    "Allow": "Izinkan",
    "Don't allow": "Jangan izinkan",
    "More actions": "Tindakan lainnya",
    "Compare packages": "Bandingkan paket",
    "Let's customize your experience": "Mari sesuaikan pengalaman Anda",
    "Watch now": "Tonton sekarang",
    "Sign in with remote": "Masuk dengan remote",

    "Subscribe": "Berlangganan",
    "Login": "Masuk",
    "Log in": "Masuk",
    "Sign in": "Masuk",
    "Sign out": "Keluar",
    "Logout": "Keluar",
    "Save": "Simpan",
    "Save Information": "Simpan Informasi",
    "Delete": "Hapus",
    "Delete All": "Hapus Semua",
    "Done": "Selesai",
    "Next": "Berikutnya",
    "Back": "Kembali",
    "Retry": "Coba Lagi",
    "Update payment": "Perbarui pembayaran",
    "Watch offline": "Tonton offline",
    "Download & Watch": "Unduh & Tonton",
    "Download & watch": "Unduh & tonton",
    "Ad-free": "Tanpa iklan",
    "Ad- free": "Tanpa iklan",
    "Parental Control": "Kontrol Orang Tua",
    "Parental control": "Kontrol orang tua",
    "Full HD": "Full HD",
    "Live TV": "TV Live",
    "Help Center": "Pusat Bantuan",
    "Contact Us": "Hubungi Kami",
    "Account Settings": "Pengaturan Akun",
    "Subscription Management": "Manajemen Langganan",
    "Devices Management": "Manajemen Perangkat",
    "My List": "Daftar Saya",
    "Continue Watching": "Lanjutkan Menonton",

    "Things to consider when you try to upgrade.": "Hal yang perlu dipertimbangkan saat Anda mencoba upgrade.",
    "I subscribed but cannot watch specific content": "Saya berlangganan tetapi tidak dapat menonton konten tertentu",
    "What are the packages that include Sports contents?": "Apa saja paket yang mencakup konten olahraga?",
    "What is VIP | Sports?": "Apa itu VIP | Sports?",
    "I am unable to upgrade to the VIP | Sports package.": "Saya tidak dapat upgrade ke paket VIP | Sports.",
    "How do I stop getting advertisements?": "Bagaimana cara berhenti melihat iklan?",
    "Plans & Packages": "Paket & Langganan",
    "Will I lose my current offer if I upgrade to another package?": "Apakah saya akan kehilangan penawaran saat ini jika upgrade ke paket lain?",
    "Countries list for Mobile Log in option": "Daftar negara untuk opsi login melalui ponsel",
    "How to upgrade?": "Bagaimana cara upgrade?",
    "How to upgrade ?": "Bagaimana cara upgrade?",
}

# =========================
# DO NOT TRANSLATE EXACTLY
# =========================

DO_NOT_TRANSLATE_EXACT = {
    "NA",
    "MBC",
    "MBC Shahid",
    "Shahid Originals",
    "MBC Shahid Originals",
    "VIP",
    "VIP Sports",
    "VIP BigTime",
    "VIP Mobile",
    "Ultimate",
    "GOBOX",
    "Big Time",
    "SBA",
    "RSL",
    "Al Arabiya",
    "Al Thaqafeya",
    "Iqraa",
    "Anime",
    "Bollywood",
    "Khaliji",
    "Hero",
    "dynamicTitle",
    "Full HD",
    "Live TV",
    "Shorts",
    "Apple Pay",
    "Google Pay",
    "Google Play",
    "Google Play Store",
    "App Store",
    "iTunes",
    "Wallet & Apple Pay",
}

PROTECTED_TERMS = [
    "MBC Shahid Originals",
    "MBC Shahid",
    "Shahid Originals",
    "Google Play Store",
    "Wallet & Apple Pay",
    "Credit/Debit Card",
    "Mobile Pay",
    "Mobile Payment",
    "Apple Pay",
    "Google Pay",
    "Google Play",
    "App Store",
    "VIP Sports",
    "VIP BigTime",
    "VIP Mobile",
    "Full HD",
    "Live TV",
    "Al Arabiya",
    "Al Thaqafeya",
    "Big Time",
    "MBC",
    "VIP",
    "Ultimate",
    "GOBOX",
    "SBA",
    "RSL",
    "Iqraa",
    "Shorts",
    "iTunes",
]

# =========================
# SYMBOL / CORRUPTION FIXES
# =========================

PIPE_CORRUPTION_FIXES = {
    "114;": "|",
    "124;": "|",
    "12.4;": "|",
    "1 24;": "|",
    "1.24;": "|",
    "I24;": "|",
    "l24;": "|",
    "&#124;": "|",
    "&124;": "|",
}

BROKEN_TAG_PATTERNS = [
    r"<\s*x\s*0\s*>",
    r"<\s*/\s*x\s*0\s*>",
    r"<\s*x\s*(\d+)\s*>",
    r"<\s*/\s*x\s*(\d+)\s*>",
    r"&lt;\s*x\s*(\d+)\s*&gt;",
    r"&lt;\s*/\s*x\s*(\d+)\s*&gt;",
]

FINAL_FIXES = {
    "_ _ BRAND _ 0 _ _": "",
    "_ _ BRAND _ 1 _ _": "",
    "_ _ BRAND _ 2 _ _": "",
    "_ _ BRAND _ 3 _ _": "",
    "__ BRAND _ 0 __": "",
    "__ BRAND _ 1 __": "",
    "__ BRAND _ 2 __": "",
    "__BRAND_0__": "",
    "__BRAND_1__": "",
    "__BRAND_2__": "",

    "Anak-anak": "Anak",
    "anak-anak": "anak",

    "Aku": "Saya",
    " aku ": " saya ",
    "Anda langganan": "Anda berlangganan",
    "Saya...": "Saya ",
    "...": " ",
    "Bagaimana cara saya": "Bagaimana cara",
    "Bagaimana saya": "Bagaimana cara saya",

    "Pembayaran tidak pergi melalui": "Pembayaran tidak berhasil",
    "Pembayaran kembali tidak pergi melalui": "Percobaan pembayaran ulang tidak berhasil",
    "Membayar Apple": "Apple Pay",
    "Membayar Google": "Google Pay",
    "Dompet & Membayar Apel": "Wallet & Apple Pay",
    "Toko App": "App Store",
    "Pergi ke": "Buka",
    "Mutakhirkan pembayaran": "Perbarui pembayaran",
    "Citra profil": "Gambar profil",
    "Sandi": "Kata Sandi",
    "Kolom ini diperlukan": "Kolom ini wajib diisi",
    "Berkas ini diperlukan": "Kolom ini wajib diisi",

    "TV langsung": "TV Live",
    "Langsung": "Live",
    "Iklan asli": "Iklan Native",
    "Kualitas penuh HD": "Full HD",
    "unduh dan tonton": "unduh & tonton",
    "Unduh dan Tonton": "Unduh & Tonton",
    "Unduh & menonton": "Unduh & tonton",
    "Download & tonton": "Unduh & tonton",
    "VIP berlangganan": "langganan VIP",
    "Ponsel Pay": "Pembayaran Mobile",
    "Mobile Pay": "Pembayaran Mobile",
    "Kartu Kredit/Debit": "Kartu Kredit/Debit",
}

# =========================
# HEADER NORMALIZATION
# =========================

def normalize_header(value):
    if value is None:
        return ""

    text = str(value)
    text = text.replace("\xa0", " ")
    text = text.strip().lower()

    # Normalize underscores, hyphens, and repeated spaces
    text = text.replace("_", " ")
    text = text.replace("-", " ")
    text = re.sub(r"\s+", " ", text)

    # Normalize common typo
    text = text.replace("enlgish", "english")

    # Remove parentheses for easier matching too
    text_no_parens = text.replace("(", "").replace(")", "")
    text_no_parens = re.sub(r"\s+", " ", text_no_parens).strip()

    return text_no_parens

# =========================
# CLEANING HELPERS
# =========================

def normalize_source_text(text):
    text = str(text)
    text = text.replace("\xa0", " ")
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = text.strip()

    for wrong, correct in SOURCE_FIXES.items():
        text = text.replace(wrong, correct)

    return text


def apply_final_fixes(text):
    text = str(text)

    for wrong, correct in PIPE_CORRUPTION_FIXES.items():
        text = text.replace(wrong, correct)

    text = re.sub(r"_\s*_?\s*BRAND\s*_?\s*\d+\s*_?\s*_", "", text, flags=re.IGNORECASE)
    text = re.sub(r"__\s*BRAND\s*_\s*\d+\s*__", "", text, flags=re.IGNORECASE)

    for pattern in BROKEN_TAG_PATTERNS:
        text = re.sub(pattern, "", text, flags=re.IGNORECASE)

    for wrong, correct in FINAL_FIXES.items():
        text = text.replace(wrong, correct)

    text = re.sub(r"\s+([.,!?;:])", r"\1", text)
    text = re.sub(r"\s*\|\s*", " | ", text)
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"(?m)^(\d+)\s*\.\s*", r"\1. ", text)

    return text.strip()

# =========================
# URL / PLACEHOLDER DETECTION
# =========================

def is_real_url(value):
    value = str(value).strip()
    return bool(
        re.fullmatch(r"https?://\S+", value)
        or re.fullmatch(r"www\.\S+", value)
    )


def should_never_translate(value):
    value = str(value).strip()

    if not value:
        return True

    if value in DO_NOT_TRANSLATE_EXACT:
        return True

    if is_real_url(value):
        return True

    if re.fullmatch(r"\{[^{}]+\}", value):
        return True

    if re.fullmatch(r"\[[^\[\]]+\]", value):
        return True

    if re.fullmatch(r"%\w", value):
        return True

    return False

# =========================
# PROTECT SPECIAL PARTS
# =========================

def protect_special_parts(text):
    protected_items = []

    patterns = [
        r"https?://\S+",
        r"www\.\S+",
        r"\{[^{}]+\}",
        r"\[[^\[\]]+\]",
        r"%\w",
        r"\$\w+",
        r"\|",
    ]

    for term in sorted(PROTECTED_TERMS, key=len, reverse=True):
        patterns.append(re.escape(term))

    combined = re.compile("|".join(patterns), flags=re.IGNORECASE)

    def replacer(match):
        protected_items.append(match.group(0))
        idx = len(protected_items) - 1
        return f"<x{idx}></x{idx}>"

    protected_text = combined.sub(replacer, text)

    return protected_text, protected_items


def restore_special_parts(text, protected_items):
    restored = str(text)

    for idx, original in enumerate(protected_items):
        original = unescape(original)

        possible_patterns = [
            f"<x{idx}></x{idx}>",
            f"<x{idx}> </x{idx}>",
            f"< x{idx} ></ x{idx} >",
            f"< x{idx} > < / x{idx} >",
            f"&lt;x{idx}&gt;&lt;/x{idx}&gt;",
            f"&lt;x{idx}&gt; &lt;/x{idx}&gt;",
            f"&lt; x{idx} &gt;&lt; / x{idx} &gt;",
            f"&lt; x{idx} &gt; &lt; / x{idx} &gt;",
        ]

        for pattern in possible_patterns:
            restored = restored.replace(pattern, original)

        restored = re.sub(
            rf"<\s*x\s*{idx}\s*>.*?<\s*/\s*x\s*{idx}\s*>",
            original,
            restored,
            flags=re.IGNORECASE,
        )

        restored = re.sub(
            rf"&lt;\s*x\s*{idx}\s*&gt;.*?&lt;\s*/\s*x\s*{idx}\s*&gt;",
            original,
            restored,
            flags=re.IGNORECASE,
        )

    restored = unescape(restored)

    restored = re.sub(r"<\s*/?\s*x\s*\d+\s*>", "", restored, flags=re.IGNORECASE)
    restored = re.sub(r"&lt;\s*/?\s*x\s*\d+\s*&gt;", "", restored, flags=re.IGNORECASE)

    return restored

# =========================
# NUMBERED LIST HANDLING
# =========================

def split_numbered_prefix(line):
    match = re.match(r"^(\s*\d+\s*[\.\)]\s+)(.+)$", line)

    if not match:
        return "", line

    return match.group(1), match.group(2)

# =========================
# TRANSLATION
# =========================

TRANSLATION_CACHE = {}

def translate_segment(segment):
    segment = str(segment)

    if not segment.strip():
        return segment

    if should_never_translate(segment):
        return segment.strip()

    normalized = normalize_source_text(segment)

    if normalized in MANUAL_TRANSLATIONS:
        return MANUAL_TRANSLATIONS[normalized]

    if normalized in DO_NOT_TRANSLATE_EXACT:
        return normalized

    if normalized in TRANSLATION_CACHE:
        return TRANSLATION_CACHE[normalized]

    protected_text, protected_items = protect_special_parts(normalized)

    translated = argostranslate.translate.translate(
        protected_text,
        FROM_LANG,
        TO_LANG
    )

    translated = restore_special_parts(translated, protected_items)
    translated = apply_final_fixes(translated)

    TRANSLATION_CACHE[normalized] = translated

    return translated


def translate_text(text):
    original = str(text).strip()

    # Empty
    if not original:
        return ""

    # Cells containing only ""
    if original in {'""', "''"}:
        return original

    # Cells containing only spaces inside quotes
    if re.fullmatch(r'"\s*"', original):
        return '""'

    if SKIP_FORMULAS and original.startswith("="):
        return original

    normalized = normalize_source_text(original)

    if should_never_translate(normalized):
        return normalized

    if normalized in MANUAL_TRANSLATIONS:
        return MANUAL_TRANSLATIONS[normalized]

    lines = normalized.split("\n")
    translated_lines = []

    for line in lines:
        if not line.strip():
            translated_lines.append("")
            continue

        prefix, body = split_numbered_prefix(line)
        translated_body = translate_segment(body)
        translated_lines.append(prefix + translated_body)

    result = "\n".join(translated_lines)
    result = apply_final_fixes(result)

    return result

# =========================
# TRANSLATABLE SOURCE CHECK
# =========================

def is_translatable_source(value):
    if value is None:
        return False

    if isinstance(value, (int, float, bool)):
        return False

    value = str(value).strip()

    if not value:
        return False

    # Skip "" completely
    if value in {'""', "''"}:
        return False

    # Skip "   "
    if re.fullmatch(r'"\s*"', value):
        return False

    if SKIP_FORMULAS and value.startswith("="):
        return False

    if should_never_translate(value):
        return False

    return bool(re.search(r"[A-Za-z]", value))

# =========================
# COLUMN DETECTION
# =========================

def find_nearest_ai_sub_column_to_right(ws, source_col_idx):
    """
    Finds the nearest AI Sub / AI Subs column to the right of a source column.
    It stops at the first matching target header.
    """
    for col_idx in range(source_col_idx + 1, ws.max_column + 1):
        header_value = ws.cell(row=HEADER_ROW, column=col_idx).value
        normalized = normalize_header(header_value)

        if normalized in TARGET_AI_HEADERS:
            return col_idx

    return None


def find_source_target_pairs(ws):
    """
    Finds all source columns whose row-1 header matches SOURCE_HEADERS_TO_TRANSLATE,
    then pairs each with the nearest AI Sub / AI Subs column to its right.
    """
    pairs = []

    normalized_source_headers = {normalize_header(h) for h in SOURCE_HEADERS_TO_TRANSLATE}
    normalized_target_headers = {normalize_header(h) for h in TARGET_AI_HEADERS}

    for col_idx in range(1, ws.max_column + 1):
        header_value = ws.cell(row=HEADER_ROW, column=col_idx).value
        normalized_header = normalize_header(header_value)

        if normalized_header not in normalized_source_headers:
            continue

        target_col_idx = None

        for right_col_idx in range(col_idx + 1, ws.max_column + 1):
            right_header_value = ws.cell(row=HEADER_ROW, column=right_col_idx).value
            right_normalized_header = normalize_header(right_header_value)

            if right_normalized_header in normalized_target_headers:
                target_col_idx = right_col_idx
                break

        if target_col_idx:
            pairs.append((col_idx, target_col_idx))
        else:
            source_letter = openpyxl.utils.get_column_letter(col_idx)
            print(f"  Warning: source column {source_letter} has no AI Sub / AI Subs column to its right.")

    return pairs

# =========================
# INSTALL ARGOS MODEL
# =========================

print("Preparing translation model...")

argostranslate.package.update_package_index()
available_packages = argostranslate.package.get_available_packages()

package = next(
    p for p in available_packages
    if p.from_code == FROM_LANG and p.to_code == TO_LANG
)

download_path = package.download()
argostranslate.package.install_from_path(download_path)

print("Translation model ready.")

# =========================
# PROCESS WORKBOOK
# =========================

if not INPUT_FILE.exists():
    raise FileNotFoundError(f"Input file not found: {INPUT_FILE}")

print(f"Processing workbook: {INPUT_FILE.name}")

wb = openpyxl.load_workbook(INPUT_FILE)

worksheets = wb.worksheets

if SKIP_FIRST_SHEET and worksheets:
    worksheets_to_process = worksheets[1:]
    print(f"Skipping first sheet: {worksheets[0].title}")
else:
    worksheets_to_process = worksheets

total_pairs_found = 0
total_updated = 0
total_skipped_existing = 0
total_skipped_non_text = 0
total_sheets_without_pairs = 0

for ws in worksheets_to_process:
    print(f"\nSheet: {ws.title}")

    pairs = find_source_target_pairs(ws)

    if not pairs:
        print("  No matching source columns with AI Sub / AI Subs target columns found.")
        total_sheets_without_pairs += 1
        continue

    total_pairs_found += len(pairs)

    for source_col_idx, target_col_idx in pairs:
        source_letter = openpyxl.utils.get_column_letter(source_col_idx)
        target_letter = openpyxl.utils.get_column_letter(target_col_idx)

        source_header = ws.cell(row=HEADER_ROW, column=source_col_idx).value
        target_header = ws.cell(row=HEADER_ROW, column=target_col_idx).value

        print(f"  Translating {source_letter} ({source_header}) -> {target_letter} ({target_header})")

        updated_in_pair = 0
        skipped_existing_in_pair = 0
        skipped_non_text_in_pair = 0

        for row_idx in tqdm(range(HEADER_ROW + 1, ws.max_row + 1), desc=f"{ws.title}: {source_header}", unit="rows",):
            source_cell = ws.cell(row=row_idx, column=source_col_idx)
            target_cell = ws.cell(row=row_idx, column=target_col_idx)

            source_value = source_cell.value

            # Preserve empty quoted strings without translating
            if isinstance(source_value, str):
                stripped = source_value.strip()

                if (
                    stripped in {'""', "''"}
                    or re.fullmatch(r'"\s*"', stripped)
                ):
                    target_cell.value = '""'
                    skipped_non_text_in_pair += 1
                    continue

            if not is_translatable_source(source_value):
                skipped_non_text_in_pair += 1
                continue

            if not OVERWRITE_EXISTING_AI_SUB:
                existing_target = target_cell.value
                if existing_target is not None and str(existing_target).strip():
                    skipped_existing_in_pair += 1
                    continue

            try:
                translated = translate_text(source_value)
                if (
                    isinstance(source_value, str)
                    and source_value.startswith('"')
                    and source_value.endswith('"')
                ):
                    translated = f'"{translated}"'

                target_cell.value = translated 
                updated_in_pair += 1
            except Exception as e:
                print(f"    Error at {ws.title}!{source_cell.coordinate}: {e}")

        total_updated += updated_in_pair
        total_skipped_existing += skipped_existing_in_pair
        total_skipped_non_text += skipped_non_text_in_pair

        print(f"    Updated: {updated_in_pair}")
        print(f"    Skipped existing: {skipped_existing_in_pair}")
        print(f"    Skipped empty/non-text: {skipped_non_text_in_pair}")

wb.save(OUTPUT_FILE)

print("\nDone.")
print(f"Total sheets processed: {len(worksheets_to_process)}")
print(f"Total sheets without matching pairs: {total_sheets_without_pairs}")
print(f"Total source-target pairs found: {total_pairs_found}")
print(f"Total translations updated: {total_updated}")
print(f"Total skipped existing translations: {total_skipped_existing}")
print(f"Total skipped empty/non-text cells: {total_skipped_non_text}")
print(f"Saved workbook to: {OUTPUT_FILE}")